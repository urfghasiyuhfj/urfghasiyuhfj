#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PPM 系统测试数据生成器
生成可疑物料、供货量、供应商 PPM 的测试数据
"""

import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 配置参数
RECORD_COUNT_SUSPICIOUS = 5000  # 可疑物料记录数
RECORD_COUNT_SUPPLY = 2000      # 供货量记录数
RECORD_COUNT_PPM = 1000         # 供应商 PPM 记录数

# 基础数据
PLANTS = ['河西工厂', '宝骏基地', '青岛工厂', '重庆工厂', '柳州总部']
BASE_CODES = ['河西', '宝骏', '青岛', '重庆']
FAULT_TYPES = ['外观缺陷', '尺寸超差', '性能不良', '装配问题', '材料问题', '包装破损', '标签错误']
FAILURE_DESCS = ['表面划痕', '尺寸偏大', '硬度不足', '装配间隙大', '色差', '变形', '锈蚀', '断裂', '磨损', '松动']
FUNCTION_MODULES = ['发动机', '变速箱', '底盘', '车身', '电气', '内饰', '外饰']
BRANDS = ['五菱', '宝骏', '雪佛兰', '别克']
SHIFT_SECTIONS = ['早班', '中班', '晚班', 'A工段', 'B工段', 'C工段']
PROD_AREAS = ['冲压车间', '焊装车间', '涂装车间', '总装车间', '发动机车间']

# 供应商数据池
SUPPLIERS = [
    ('S001', '上海汽车零部件有限公司'),
    ('S002', '北京精密制造股份有限公司'),
    ('S003', '广州橡胶制品厂'),
    ('S004', '深圳电子科技有限公司'),
    ('S005', '武汉钢铁加工有限公司'),
    ('S006', '重庆塑料制品厂'),
    ('S007', '成都机械加工厂'),
    ('S008', '天津五金制品有限公司'),
    ('S009', '苏州精密模具厂'),
    ('S010', '杭州轴承制造厂'),
    ('S011', '南京弹簧有限公司'),
    ('S012', '青岛密封件厂'),
    ('S013', '沈阳铸造有限公司'),
    ('S014', '大连锻造厂'),
    ('S015', '宁波紧固件有限公司'),
    ('S016', '温州标准件厂'),
    ('S017', '台州齿轮制造厂'),
    ('S018', '嘉兴粉末冶金厂'),
    ('S019', '绍兴冲压件厂'),
    ('S020', '金华注塑件厂'),
]

# 零件数据池
PARTS = [
    ('P001', '发动机支架'),
    ('P002', '变速箱壳体'),
    ('P003', '转向节'),
    ('P004', '制动盘'),
    ('P005', '减震器'),
    ('P006', '轮毂轴承'),
    ('P007', '传动轴'),
    ('P008', '排气管'),
    ('P009', '油箱'),
    ('P010', '座椅骨架'),
    ('P011', '仪表板'),
    ('P012', '保险杠'),
    ('P013', '车门内板'),
    ('P014', '发动机盖'),
    ('P015', '行李箱盖'),
    ('P016', '前挡风玻璃'),
    ('P017', '后视镜'),
    ('P018', '大灯总成'),
    ('P019', '尾灯总成'),
    ('P020', '轮胎'),
    ('P021', '蓄电池'),
    ('P022', '发电机'),
    ('P023', '启动机'),
    ('P024', '空调压缩机'),
    ('P025', '水泵'),
    ('P026', '机油泵'),
    ('P027', '燃油泵'),
    ('P028', '火花塞'),
    ('P029', '滤清器'),
    ('P030', '皮带轮'),
]


def random_date(start_date, end_date):
    """生成随机日期"""
    time_between = end_date - start_date
    days_between = time_between.days
    random_days = random.randrange(days_between)
    return start_date + datetime.timedelta(days=random_days)


def generate_suspicious_material_data():
    """生成可疑物料测试数据"""
    print(f"正在生成 {RECORD_COUNT_SUSPICIOUS} 条可疑物料数据...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "可疑物料统计"
    
    # 表头
    headers = [
        "区域工厂", "录入日期", "录入人员", "零件号", "零件名称", "功能模块",
        "供应商名称", "供应商代码", "车型|机型", "失效描述", "故障类别", "数量",
        "开单日期", "班次|工段", "产生区域", "是否供应商责任", "备注", "供应商&零件",
        "可疑物料数量", "供货量", "品牌"
    ]
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 生成数据
    start_date = datetime.date(2024, 1, 1)
    end_date = datetime.date(2025, 3, 15)
    
    for i in range(RECORD_COUNT_SUSPICIOUS):
        supplier_code, supplier_name = random.choice(SUPPLIERS)
        part_code, part_name = random.choice(PARTS)
        plant = random.choice(PLANTS)
        record_date = random_date(start_date, end_date)
        order_date = random_date(start_date, record_date)
        
        quantity = random.randint(1, 100)
        defect_count = random.randint(1, quantity)
        supply_qty = random.randint(100, 10000)
        
        row = [
            plant,                                          # 区域工厂
            record_date.strftime('%Y-%m-%d'),              # 录入日期
            f"录入员{random.randint(1, 20)}",               # 录入人员
            part_code,                                      # 零件号
            part_name,                                      # 零件名称
            random.choice(FUNCTION_MODULES),                # 功能模块
            supplier_name,                                  # 供应商名称
            supplier_code,                                  # 供应商代码
            f"车型{random.randint(1, 10)}",                 # 车型|机型
            random.choice(FAILURE_DESCS),                   # 失效描述
            random.choice(FAULT_TYPES),                     # 故障类别
            quantity,                                       # 数量
            order_date.strftime('%Y-%m-%d'),               # 开单日期
            random.choice(SHIFT_SECTIONS),                  # 班次|工段
            random.choice(PROD_AREAS),                      # 产生区域
            random.choice(['Y', 'N']),                      # 是否供应商责任
            f"备注信息{i+1}",                               # 备注
            f"{supplier_code}&{part_code}",                # 供应商&零件
            defect_count,                                   # 可疑物料数量
            supply_qty,                                     # 供货量
            random.choice(BRANDS),                          # 品牌
        ]
        ws.append(row)
        
        if (i + 1) % 1000 == 0:
            print(f"  已生成 {i + 1} 条...")
    
    filename = "可疑物料测试数据.xlsx"
    wb.save(filename)
    print(f"✓ 可疑物料数据已保存: {filename}")
    return filename


def generate_supply_volume_data():
    """生成供货量测试数据"""
    print(f"\n正在生成 {RECORD_COUNT_SUPPLY} 条供货量数据...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "供货量数据"
    
    # 表头
    headers = [
        "fiscal_year", "base_code", "supplier_code", "supplier_name",
        "part_code", "part_name", "plant_id",
        "month_1_no", "month_2_no", "month_3_no", "month_4_no",
        "month_5_no", "month_6_no", "month_7_no", "month_8_no",
        "month_9_no", "month_10_no", "month_11_no", "month_12_no",
        "supplier_code&part_code"
    ]
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 生成数据
    fiscal_year = 2025
    
    for i in range(RECORD_COUNT_SUPPLY):
        supplier_code, supplier_name = random.choice(SUPPLIERS)
        part_code, part_name = random.choice(PARTS)
        base_code = random.choice(BASE_CODES)
        
        # 生成12个月的供货量数据
        month_data = [random.randint(0, 5000) for _ in range(12)]
        
        row = [
            fiscal_year,                                    # fiscal_year
            base_code,                                      # base_code
            supplier_code,                                  # supplier_code
            supplier_name,                                  # supplier_name
            part_code,                                      # part_code
            part_name,                                      # part_name
            f"{base_code}{random.randint(1, 5)}厂",         # plant_id
        ] + month_data + [
            f"{supplier_code}&{part_code}",                # supplier_code&part_code
        ]
        ws.append(row)
        
        if (i + 1) % 500 == 0:
            print(f"  已生成 {i + 1} 条...")
    
    filename = "供货量测试数据.xlsx"
    wb.save(filename)
    print(f"✓ 供货量数据已保存: {filename}")
    return filename


def generate_supplier_ppm_data():
    """生成供应商 PPM 测试数据"""
    print(f"\n正在生成 {RECORD_COUNT_PPM} 条供应商 PPM 数据...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "供应商PPM"
    
    # 表头
    headers = [
        "PPM月份", "供应商编码", "供应商名称", "月不合格数", "月供货量", "供应商总PPM",
        "河西可疑物料", "河西供货量", "河西供应商PPM",
        "宝骏可疑物料", "宝骏供货量", "宝骏供应商PPM",
        "青岛可疑物料", "青岛供货量", "青岛供应商PPM",
        "重庆可疑物料", "重庆供货量", "重庆供应商PPM"
    ]
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 生成数据
    months = ['202401', '202402', '202403', '202404', '202405', '202406',
              '202407', '202408', '202409', '202410', '202411', '202412',
              '202501', '202502', '202503']
    
    for i in range(RECORD_COUNT_PPM):
        supplier_code, supplier_name = random.choice(SUPPLIERS)
        ppm_month = random.choice(months)
        
        # 计算总不合格数和总供货量
        month_defect = random.randint(0, 500)
        month_supply = random.randint(1000, 50000)
        total_ppm = round((month_defect / month_supply) * 1000000, 2) if month_supply > 0 else 0
        
        # 各基地数据
        base_data = []
        for _ in range(4):  # 4个基地
            defect = random.randint(0, 200)
            supply = random.randint(100, 15000)
            ppm = round((defect / supply) * 1000000, 2) if supply > 0 else 0
            base_data.extend([defect, supply, ppm])
        
        row = [
            ppm_month,          # PPM月份
            supplier_code,      # 供应商编码
            supplier_name,      # 供应商名称
            month_defect,       # 月不合格数
            month_supply,       # 月供货量
            total_ppm,          # 供应商总PPM
        ] + base_data
        
        ws.append(row)
        
        if (i + 1) % 500 == 0:
            print(f"  已生成 {i + 1} 条...")
    
    filename = "供应商PPM测试数据.xlsx"
    wb.save(filename)
    print(f"✓ 供应商 PPM 数据已保存: {filename}")
    return filename


def main():
    print("=" * 60)
    print("PPM 系统测试数据生成器")
    print("=" * 60)
    
    try:
        # 生成三类测试数据
        suspicious_file = generate_suspicious_material_data()
        supply_file = generate_supply_volume_data()
        ppm_file = generate_supplier_ppm_data()
        
        print("\n" + "=" * 60)
        print("测试数据生成完成！")
        print("=" * 60)
        print(f"\n生成的文件:")
        print(f"  1. {suspicious_file}")
        print(f"  2. {supply_file}")
        print(f"  3. {ppm_file}")
        print(f"\n请通过前端页面【数据导入】功能导入这些文件")
        print("导入顺序建议: 供应商 PPM → 供货量 → 可疑物料")
        
    except ImportError as e:
        print(f"\n错误: 缺少必要的库 - {e}")
        print("请安装 openpyxl: pip install openpyxl")
    except Exception as e:
        print(f"\n错误: {e}")


if __name__ == "__main__":
    main()
